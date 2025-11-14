import os
import io
import uuid
import threading
import traceback
from flask import request, jsonify, send_file, flash, redirect, url_for, current_app, abort
from flask_login import login_required, current_user
from sqlalchemy import or_, delete
from sqlalchemy.orm import selectinload

from flowork.models import db, Product, Variant, StoreStock, Setting, Store
from flowork.utils import clean_string_upper, get_choseong, generate_barcode, get_sort_key

from flowork.services.excel import (
    import_excel_file,
    export_db_to_excel,
    export_stock_check_excel,
    _process_stock_update_excel,
    verify_stock_excel
)
from flowork.services.db import sync_missing_data_in_db

from . import api_bp
from .utils import admin_required, _get_or_create_store_stock
from .tasks import TASKS, run_async_stock_upsert, run_async_import_db

@api_bp.route('/api/verify_excel', methods=['POST'])
@login_required
def verify_excel_upload():
    if 'excel_file' not in request.files:
        return jsonify({'status': 'error', 'message': '파일이 없습니다.'}), 400
    
    file = request.files['excel_file']
    stock_type = request.form.get('stock_type', 'store') 
    
    task_id = str(uuid.uuid4())
    temp_path = f"/tmp/verify_{task_id}.xlsx"
    file.save(temp_path)
    
    result = verify_stock_excel(temp_path, request.form, stock_type)
    
    if os.path.exists(temp_path):
        os.remove(temp_path)
        
    return jsonify(result)

@api_bp.route('/import_excel', methods=['POST'])
@admin_required
def import_excel():
    if not current_user.brand_id or current_user.store_id:
        abort(403, description="상품 DB 임포트는 본사 관리자만 가능합니다.")
        
    file = request.files.get('excel_file')
    if not file:
        return jsonify({'status': 'error', 'message': '파일이 없습니다.'}), 400

    task_id = str(uuid.uuid4())
    TASKS[task_id] = {'status': 'processing', 'current': 0, 'total': 0, 'percent': 0}
    
    temp_filename = f"/tmp/import_{task_id}.xlsx"
    file.save(temp_filename)
    
    current_brand_id = current_user.current_brand_id
    
    thread = threading.Thread(
        target=run_async_import_db,
        args=(
            current_app._get_current_object(), 
            task_id, 
            temp_filename, 
            request.form, 
            current_brand_id
        )
    )
    thread.start()
    
    return jsonify({'status': 'success', 'task_id': task_id, 'message': '업로드 작업을 시작했습니다.'})

@api_bp.route('/export_db_excel')
@login_required
def export_db_excel():
    if current_user.is_super_admin:
         abort(403, description="슈퍼 관리자는 이 API를 사용할 수 없습니다.")

    output, download_name, error_message = export_db_to_excel(current_user.current_brand_id)
    if error_message:
        flash(error_message, 'warning')
        return redirect(url_for('ui.setting_page'))
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )

@api_bp.route('/sync_missing_data', methods=['POST'])
@login_required
def sync_missing_data():
    if not current_user.is_admin:
         abort(403, description="데이터 동기화는 관리자 계정만 사용할 수 있습니다.")

    success, message, category = sync_missing_data_in_db(current_user.current_brand_id)
    flash(message, category)
    
    if current_user.store_id:
        return redirect(url_for('ui.stock_management'))
    else:
        return redirect(url_for('ui.setting_page'))

@api_bp.route('/update_store_stock_excel', methods=['POST'])
@login_required
def update_store_stock_excel():
    if not current_user.store_id and not (current_user.is_admin and not current_user.store_id):
         abort(403, description="매장 재고 업데이트는 매장 관리자 또는 본사 관리자만 사용할 수 있습니다.")
    
    target_store_id = None
    allow_create = False

    if current_user.is_admin and not current_user.store_id:
        allow_create = True
    
    if current_user.store_id:
        target_store_id = current_user.store_id
        allow_create = False
    elif 'target_store_id' in request.form:
        target_store_id = int(request.form.get('target_store_id'))
    
    if not target_store_id:
        return jsonify({'status': 'error', 'message': '재고를 업데이트할 대상 매장을 확인할 수 없습니다.'}), 400

    file = request.files.get('excel_file')
    current_brand_id = current_user.current_brand_id
    
    excluded_str = request.form.get('excluded_row_indices', '')
    excluded_indices = [int(x) for x in excluded_str.split(',')] if excluded_str else []

    if 'col_pn' in request.form:
        task_id = str(uuid.uuid4())
        TASKS[task_id] = {'status': 'processing', 'current': 0, 'total': 0, 'percent': 0}
        
        temp_filename = f"/tmp/verify_{task_id}.xlsx"
        file.save(temp_filename)
        
        thread = threading.Thread(
            target=run_async_stock_upsert,
            args=(
                current_app._get_current_object(), 
                task_id, 
                temp_filename, 
                request.form, 
                'store', 
                current_brand_id, 
                target_store_id,
                excluded_indices,
                allow_create
            )
        )
        thread.start()
        
        return jsonify({'status': 'success', 'task_id': task_id, 'message': '업데이트 작업을 시작했습니다.'})

    elif 'barcode_col' in request.form:
        try:
            processed, created, message, category = _process_stock_update_excel(
                file, request.form, 'store', 
                current_brand_id, 
                target_store_id
            )
            flash(message, category)
            return jsonify({'status': 'success', 'message': message})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'오류: {e}'}), 500
            
    else:
        return jsonify({'status': 'error', 'message': '알 수 없는 폼 형식입니다.'}), 400


@api_bp.route('/update_hq_stock_excel', methods=['POST'])
@admin_required
def update_hq_stock_excel():
    if not current_user.brand_id or current_user.store_id:
        abort(403, description="본사 재고 업데이트는 본사 관리자만 가능합니다.")

    file = request.files.get('excel_file')
    current_brand_id = current_user.current_brand_id
    
    excluded_str = request.form.get('excluded_row_indices', '')
    excluded_indices = [int(x) for x in excluded_str.split(',')] if excluded_str else []

    if 'col_pn' in request.form:
        task_id = str(uuid.uuid4())
        TASKS[task_id] = {'status': 'processing', 'current': 0, 'total': 0, 'percent': 0}
        
        temp_filename = f"/tmp/{task_id}.xlsx"
        file.save(temp_filename)
        
        thread = threading.Thread(
            target=run_async_stock_upsert,
            args=(
                current_app._get_current_object(), 
                task_id, 
                temp_filename, 
                request.form, 
                'hq', 
                current_brand_id, 
                None,
                excluded_indices,
                True
            )
        )
        thread.start()
        
        return jsonify({'status': 'success', 'task_id': task_id, 'message': '업데이트 작업을 시작했습니다.'})

    elif 'barcode_col' in request.form:
        try:
            processed, created, message, category = _process_stock_update_excel(
                file, request.form, 'hq', 
                current_brand_id, 
                None
            )
            flash(message, category)
            return jsonify({'status': 'success', 'message': message})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'오류: {e}'}), 500
        
    else:
        return jsonify({'status': 'error', 'message': '알 수 없는 폼 형식입니다.'}), 400

@api_bp.route('/export_stock_check')
@login_required
def export_stock_check():
    target_store_id = None
    
    if current_user.store_id:
        target_store_id = current_user.store_id
    elif current_user.is_admin: 
        target_store_id = request.args.get('target_store_id', type=int)
        
    if not target_store_id:
        abort(403, description="매장 정보를 확인할 수 없습니다.")
    
    output, download_name, error_message = export_stock_check_excel(
        target_store_id, 
        current_user.current_brand_id
    )
    
    if error_message:
        flash(error_message, 'error')
        return redirect(url_for('ui.check_page')) 

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )

@api_bp.route('/api/live_search', methods=['POST'])
@login_required
def live_search():
    if current_user.is_super_admin:
         return jsonify({'status': 'error', 'message': '슈퍼 관리자는 이 API를 사용할 수 없습니다.'}), 403
         
    data = request.json
    query_param = data.get('query', '')
    category_param = data.get('category', '전체')
    
    page = data.get('page', 1)
    per_page = data.get('per_page', 10)
    
    base_query = Product.query.options(selectinload(Product.variants)).filter(
        Product.brand_id == current_user.current_brand_id
    )
    showing_favorites = False

    is_searching = bool(query_param) or (category_param and category_param != '전체')

    if is_searching:
        if query_param:
            search_term_cleaned = clean_string_upper(query_param)
            search_like = '%' + search_term_cleaned + '%'
            base_query = base_query.filter(
                or_(
                    Product.product_number_cleaned.like(search_like),
                    Product.product_name_cleaned.like(search_like),
                    Product.product_name_choseong.like(search_like)
                )
            )

        if category_param and category_param != '전체':
            base_query = base_query.filter(Product.item_category == category_param)

        final_query = base_query.order_by(Product.release_year.desc(), Product.product_name)
    else:
        showing_favorites = True
        final_query = base_query.filter(Product.is_favorite == 1).order_by(Product.item_category, Product.product_name)

    pagination = final_query.paginate(page=page, per_page=per_page, error_out=False)
    products = pagination.items

    setting_rule = Setting.query.filter_by(brand_id=current_user.current_brand_id, key='IMAGE_NAMING_RULE').first()
    naming_rule = setting_rule.value if setting_rule else "{product_number}"

    results_list = []
    for product in products:
        pn = product.product_number.split(' ')[0]
        colors = ""
        sale_price_f = "가격정보없음"
        original_price_f = 0
        discount_f = "-"
        product_variants = product.variants 

        color = "00"
        if product_variants:
            colors_list = sorted(list(set(v.color for v in product_variants if v.color)))
            colors = ", ".join(colors_list)
            first_variant = product_variants[0]
            color = first_variant.color
            sale_price_f = f"{first_variant.sale_price:,d}원"
            original_price_f = first_variant.original_price
            if original_price_f and original_price_f > 0 and original_price_f != sale_price_f:
                discount_f = f"{int((1 - (first_variant.sale_price / original_price_f)) * 100)}%"
            else:
                discount_f = "0%"

        year = str(product.release_year) if product.release_year else ""
        if not year and len(pn) >= 5 and pn[3:5].isdigit():
             year = f"20{pn[3:5]}"

        try:
            filename = naming_rule.format(
                product_number=pn,
                color=color,
                year=year
            )
        except:
            filename = pn

        if filename.lower().endswith('.jpg'):
            image_pn = filename[:-4]
        else:
            image_pn = filename

        results_list.append({
            "product_id": product.id,
            "product_number": product.product_number,
            "product_name": product.product_name,
            "image_pn": image_pn,
            "colors": colors,
            "sale_price": sale_price_f,
            "original_price": original_price_f,
            "discount": discount_f
        })

    return jsonify({
        "status": "success",
        "products": results_list,
        "showing_favorites": showing_favorites,
        "selected_category": category_param,
        "current_page": pagination.page,
        "total_pages": pagination.pages,
        "total_items": pagination.total,
        "has_next": pagination.has_next,
        "has_prev": pagination.has_prev
    })

@api_bp.route('/reset_actual_stock', methods=['POST'])
@login_required
def reset_actual_stock():
    target_store_id = None
    
    if current_user.store_id:
        target_store_id = current_user.store_id
    elif current_user.is_admin:
        target_store_id = request.form.get('target_store_id', type=int)
        
    if not target_store_id:
        abort(403, description="초기화할 매장 정보를 확인할 수 없습니다.")

    try: 
        store_stock_ids_query = db.session.query(StoreStock.id).filter_by(store_id=target_store_id)
        
        stmt = db.update(StoreStock).where(
            StoreStock.id.in_(store_stock_ids_query)
        ).values(actual_stock=None)
        
        result = db.session.execute(stmt)
        db.session.commit()
        flash(f'실사재고 {result.rowcount}건 초기화 완료.', 'success')
    except Exception as e: 
        db.session.rollback()
        flash(f'초기화 오류: {e}', 'error')
        
    return redirect(url_for('ui.check_page', target_store_id=target_store_id if not current_user.store_id else None))

@api_bp.route('/api/analyze_excel', methods=['POST'])
@login_required
def analyze_excel():
    if 'excel_file' not in request.files:
        return jsonify({'status': 'error', 'message': '파일이 없습니다.'}), 400
    
    file = request.files.get('excel_file')
    if file.filename == '':
        return jsonify({'status': 'error', 'message': '파일이 선택되지 않았습니다.'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'status': 'error', 'message': '엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.'}), 400

    try:
        file_bytes = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        
        max_col_index = ws.max_column
        if max_col_index > 26: max_col_index = 26 
        column_letters = [get_column_letter(i) for i in range(1, max_col_index + 1)]
        
        preview_data = {}
        max_row_preview = min(6, ws.max_row + 1) 
        
        if max_row_preview <= 1:
             return jsonify({'status': 'error', 'message': '파일에 데이터가 없습니다.'}), 400

        for col_letter in column_letters:
            col_data = []
            col_index = column_index_from_string(col_letter)
            for i in range(1, max_row_preview):
                cell_val = ws.cell(row=i, column=col_index).value
                col_data.append(str(cell_val) if cell_val is not None else '')
            preview_data[col_letter] = col_data
            
        return jsonify({
            'status': 'success',
            'column_letters': column_letters,
            'preview_data': preview_data
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'엑셀 파일 분석 중 오류 발생: {e}'}), 500

@api_bp.route('/bulk_update_actual_stock', methods=['POST'])
@login_required
def bulk_update_actual_stock():
    data = request.json
    target_store_id = None
    
    if current_user.store_id:
        target_store_id = current_user.store_id
    elif current_user.is_admin:
        target_store_id = data.get('target_store_id')
        
    if not target_store_id:
        return jsonify({'status': 'error', 'message': '매장 정보를 확인할 수 없습니다.'}), 400
        
    items = data.get('items', [])
    if not items: 
        return jsonify({'status': 'error', 'message': '전송 상품 없음.'}), 400
    
    try:
        updated = 0
        unknown = []
        barcode_map = {clean_string_upper(item.get('barcode', '')): int(item.get('quantity', 0)) for item in items if item.get('barcode')}
        
        if not barcode_map:
            return jsonify({'status': 'error', 'message': '유효한 바코드 없음.'}), 400

        variants = db.session.query(Variant).join(Product).filter(
            Product.brand_id == current_user.current_brand_id,
            Variant.barcode_cleaned.in_(barcode_map.keys())
        ).all()
        
        variant_id_map = {v.barcode_cleaned: v.id for v in variants}
        found_barcodes = set(variant_id_map.keys())
        unknown = [b for b in barcode_map.keys() if b not in found_barcodes]
        
        if not variant_id_map:
            return jsonify({'status': 'error', 'message': 'DB에 일치하는 상품이 없습니다.'}), 404

        existing_stock = db.session.query(StoreStock).filter(
            StoreStock.store_id == target_store_id,
            StoreStock.variant_id.in_(variant_id_map.values())
        ).all()
        
        stock_map = {s.variant_id: s for s in existing_stock}
        
        new_stock_entries = []
        for barcode_cleaned, variant_id in variant_id_map.items():
            new_actual_qty = barcode_map[barcode_cleaned]
            
            if variant_id in stock_map:
                stock_map[variant_id].actual_stock = new_actual_qty
                updated += 1
            else:
                new_stock = StoreStock(
                    store_id=target_store_id,
                    variant_id=variant_id,
                    quantity=0,
                    actual_stock=new_actual_qty
                )
                new_stock_entries.append(new_stock)
                updated += 1

        if new_stock_entries:
            db.session.add_all(new_stock_entries)
            
        db.session.commit()
        msg = f"목록 {len(items)}개 항목 (SKU {updated}개) 실사재고 업데이트 완료."
        if unknown: 
            flash(f"DB에 없는 바코드 {len(unknown)}개: {', '.join(unknown[:5])}...", 'warning')
        flash(msg, 'success')
        return jsonify({'status': 'success', 'message': msg})
    except Exception as e: 
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/fetch_variant', methods=['POST'])
@login_required
def api_fetch_variant():
    data = request.json
    target_store_id = None
    
    if current_user.store_id:
        target_store_id = current_user.store_id
    elif current_user.is_admin:
        target_store_id = data.get('target_store_id')
        
    if not target_store_id:
        return jsonify({'status': 'error', 'message': '매장 정보를 확인할 수 없습니다.'}), 400

    barcode = data.get('barcode', '')
    if not barcode: 
        return jsonify({'status': 'error', 'message': '바코드 없음.'}), 400

    cleaned_barcode = clean_string_upper(barcode)
    if not cleaned_barcode:
        return jsonify({'status': 'error', 'message': f'"{barcode}" 검색 실패.'}), 404

    result = db.session.query(Variant, Product).join(Product, Variant.product_id == Product.id).filter(
        Variant.barcode_cleaned == cleaned_barcode,
        Product.brand_id == current_user.current_brand_id
    ).first()

    if result: 
        v, p = result
        
        stock = db.session.query(StoreStock).filter_by(
            variant_id=v.id,
            store_id=target_store_id
        ).first()
        
        current_stock_qty = stock.quantity if stock else 0
        
        return jsonify({
            'status': 'success', 
            'barcode': v.barcode, 
            'variant_id': v.id, 
            'product_number': p.product_number, 
            'product_name': p.product_name, 
            'color': v.color, 
            'size': v.size, 
            'sale_price': v.sale_price, 
            'store_stock': current_stock_qty
        })
    else: 
        return jsonify({'status': 'error', 'message': f'"{barcode}" 상품 없음.'}), 404

@api_bp.route('/api/search_product_by_prefix', methods=['POST'])
@login_required
def search_product_by_prefix():
    if current_user.is_super_admin:
         return jsonify({'status': 'error', 'message': '슈퍼 관리자는 이 API를 사용할 수 없습니다.'}), 403

    data = request.json
    barcode_prefix = data.get('prefix', '')

    if not barcode_prefix or len(barcode_prefix) != 11:
        return jsonify({'status': 'error', 'message': '잘못된 바코드 접두사입니다.'}), 400

    search_prefix_cleaned = clean_string_upper(barcode_prefix)

    results = Product.query.filter(
        Product.brand_id == current_user.current_brand_id,
        Product.product_number_cleaned.startswith(search_prefix_cleaned)
    ).all()

    if len(results) == 1:
        return jsonify({'status': 'success', 'product_number': results[0].product_number})
    elif len(results) > 1:
        return jsonify({'status': 'found_many', 'query': barcode_prefix})
    else:
        return jsonify({'status': 'error', 'message': f'"{barcode_prefix}"(으)로 시작하는 품번을 찾을 수 없습니다.'}), 404

@api_bp.route('/update_stock', methods=['POST'])
@login_required
def update_stock():
    data = request.json
    target_store_id = None
    
    if current_user.store_id:
        target_store_id = current_user.store_id
    elif current_user.is_admin:
        target_store_id = data.get('target_store_id')
        
    if not target_store_id:
        return jsonify({'status': 'error', 'message': '매장 정보를 확인할 수 없습니다.'}), 400

    barcode = data.get('barcode')
    change = data.get('change')
    if not barcode or change is None: 
        return jsonify({'status': 'error', 'message': '필수 데이터 누락.'}), 400
    try:
        change = int(change)
        assert change in [1, -1]

        cleaned_barcode = clean_string_upper(barcode)
        
        variant = db.session.query(Variant).join(Product).filter(
            Variant.barcode_cleaned == cleaned_barcode,
            Product.brand_id == current_user.current_brand_id
        ).first()
        
        if variant is None:
            return jsonify({'status': 'error', 'message': '상품(바코드) 없음.'}), 404
        
        stock = _get_or_create_store_stock(variant.id, target_store_id)
        
        new_stock = max(0, stock.quantity + change)
        stock.quantity = new_stock
        db.session.commit()
        
        diff = new_stock - stock.actual_stock if stock.actual_stock is not None else None
        return jsonify({
            'status': 'success', 
            'new_quantity': new_stock, 
            'barcode': barcode, 
            'new_stock_diff': diff if diff is not None else ''
        })
    except Exception as e: 
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/toggle_favorite', methods=['POST'])
@login_required
def toggle_favorite():
    if current_user.is_super_admin:
         return jsonify({'status': 'error', 'message': '슈퍼 관리자는 이 API를 사용할 수 없습니다.'}), 403

    data = request.json
    product_id = data.get('product_id')
    if not product_id: 
        return jsonify({'status': 'error', 'message': '상품 ID 없음.'}), 400
    try:
        product = Product.query.filter_by(
            id=product_id,
            brand_id=current_user.current_brand_id
        ).first()
        
        if product is None: 
            return jsonify({'status': 'error', 'message': '상품 없음.'}), 404
        
        product.is_favorite = 1 - (product.is_favorite or 0)
        new_status = product.is_favorite
        db.session.commit()
        return jsonify({'status': 'success', 'new_favorite_status': new_status})
    except Exception as e: 
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/update_actual_stock', methods=['POST'])
@login_required
def update_actual_stock():
    data = request.json
    target_store_id = None
    
    if current_user.store_id:
        target_store_id = current_user.store_id
    elif current_user.is_admin:
        target_store_id = data.get('target_store_id')
        
    if not target_store_id:
        return jsonify({'status': 'error', 'message': '매장 정보를 확인할 수 없습니다.'}), 400

    barcode = data.get('barcode')
    actual_str = data.get('actual_stock')
    if not barcode: 
        return jsonify({'status': 'error', 'message': '바코드 누락.'}), 400
    try:
        actual = int(actual_str) if actual_str and actual_str.isdigit() else None
        if actual is not None and actual < 0: 
            actual = 0

        cleaned_barcode = clean_string_upper(barcode)
        
        variant = db.session.query(Variant).join(Product).filter(
            Variant.barcode_cleaned == cleaned_barcode,
            Product.brand_id == current_user.current_brand_id
        ).first()

        if variant is None:
            return jsonify({'status': 'error', 'message': '상품(바코드) 없음.'}), 404

        stock = _get_or_create_store_stock(variant.id, target_store_id)
        
        stock.actual_stock = actual
        db.session.commit()
        
        diff = stock.quantity - actual if actual is not None else None
        return jsonify({ 
            'status': 'success', 
            'barcode': barcode, 
            'new_actual_stock': actual if actual is not None else '', 
            'new_stock_diff': diff if diff is not None else '' 
        })
    except Exception as e: 
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/update_product_details', methods=['POST'])
@admin_required
def api_update_product_details():
    if current_user.store_id or current_user.is_super_admin:
         abort(403, description="상품 정보 수정은 본사 관리자만 가능합니다.")

    data = request.json
    product_id = data.get('product_id')
    if not product_id:
        return jsonify({'status': 'error', 'message': '상품 ID 누락'}), 400

    try:
        settings_query = Setting.query.filter_by(brand_id=current_user.current_brand_id).all()
        brand_settings = {s.key: s.value for s in settings_query}

        product = Product.query.filter_by(
            id=product_id,
            brand_id=current_user.current_brand_id
        ).first()
        
        if not product:
            return jsonify({'status': 'error', 'message': '상품을 찾을 수 없음'}), 404

        product.product_name = data.get('product_name', product.product_name)
        product.product_name_cleaned = clean_string_upper(product.product_name)
        product.product_name_choseong = get_choseong(product.product_name)
        try:
            year_val = data.get('release_year')
            product.release_year = int(year_val) if year_val else None
        except (ValueError, TypeError):
            product.release_year = None
        product.item_category = data.get('item_category', product.item_category)

        variants_data = data.get('variants', [])
        variant_ids_to_delete = []
        variants_to_add = []
        variants_to_update = {}

        for v_data in variants_data:
            action = v_data.get('action')
            variant_id = v_data.get('variant_id')

            if action == 'delete' and variant_id:
                variant_ids_to_delete.append(variant_id)
            elif action == 'add':
                variant_row = {
                    'product_number': product.product_number,
                    'color': v_data.get('color'),
                    'size': v_data.get('size'),
                }
                new_barcode = generate_barcode(variant_row, brand_settings)
                if not new_barcode:
                    raise ValueError(f"새 Variant 바코드 생성 실패: {variant_row}")
                
                variants_to_add.append(Variant(
                    barcode=new_barcode,
                    product_id=product.id,
                    color=variant_row['color'],
                    size=variant_row['size'],
                    original_price=int(v_data.get('original_price', 0)),
                    sale_price=int(v_data.get('sale_price', 0)),
                    hq_quantity=0,
                    barcode_cleaned=clean_string_upper(new_barcode),
                    color_cleaned=clean_string_upper(variant_row['color']),
                    size_cleaned=clean_string_upper(variant_row['size'])
                ))
            elif action == 'update' and variant_id:
                variants_to_update[variant_id] = {
                    'color': v_data.get('color'),
                    'size': v_data.get('size'),
                    'original_price': int(v_data.get('original_price', 0)),
                    'sale_price': int(v_data.get('sale_price', 0)),
                    'color_cleaned': clean_string_upper(v_data.get('color')),
                    'size_cleaned': clean_string_upper(v_data.get('size'))
                }

        if variant_ids_to_delete:
             db.session.execute(delete(StoreStock).where(
                 StoreStock.variant_id.in_(variant_ids_to_delete)
             ))
             db.session.execute(delete(Variant).where(
                 Variant.id.in_(variant_ids_to_delete),
                 Variant.product_id == product.id 
             ))

        if variants_to_update:
            existing_variants = Variant.query.filter(
                Variant.id.in_(variants_to_update.keys()),
                Variant.product_id == product.id
            ).all()
            for variant in existing_variants:
                updates = variants_to_update.get(variant.id)
                if updates:
                    variant.color = updates['color']
                    variant.size = updates['size']
                    variant.original_price = updates['original_price']
                    variant.sale_price = updates['sale_price']
                    variant.color_cleaned = updates['color_cleaned']
                    variant.size_cleaned = updates['size_cleaned']

        if variants_to_add:
            db.session.add_all(variants_to_add)
        
        db.session.flush()
        db.session.commit()
        return jsonify({'status': 'success', 'message': '상품 정보가 업데이트되었습니다.'})

    except ValueError as ve:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'입력 값 오류: {ve}'}), 400
    except exc.IntegrityError as ie:
         db.session.rollback()
         return jsonify({'status': 'error', 'message': f'데이터베이스 오류 (바코드 중복 등): {ie.orig}'}), 400
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/product/delete/<int:product_id>', methods=['POST'])
@admin_required
def api_delete_product(product_id):
    if current_user.store_id or current_user.is_super_admin:
        abort(403, description="상품 삭제는 본사 관리자만 가능합니다.")
    
    try:
        product = Product.query.filter_by(
            id=product_id,
            brand_id=current_user.current_brand_id
        ).first()
        
        if not product:
            return jsonify({'status': 'error', 'message': '상품을 찾을 수 없음'}), 404

        product_name = product.product_name
        
        db.session.delete(product)
        db.session.commit()
        
        flash(f"상품 '{product_name}'(ID: {product_id}) 및 하위 옵션/재고가 모두 삭제되었습니다.", 'success')
        
        return redirect(url_for('ui.search_page'))

    except exc.IntegrityError as ie:
         db.session.rollback()
         flash(f"삭제 실패. 이 상품을 참조하는 다른 데이터(예: 주문 내역)가 있어 삭제할 수 없습니다. (오류: {ie.orig})", 'error')
         return redirect(url_for('ui.product_detail', product_id=product_id))
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        flash(f"서버 오류로 상품을 삭제하지 못했습니다: {e}", 'error')
        return redirect(url_for('ui.product_detail', product_id=product_id))

@api_bp.route('/api/find_product_details', methods=['POST'])
@login_required
def api_find_product_details():
    if not current_user.store_id:
        abort(403, description="이 API는 매장 계정만 사용할 수 있습니다.")

    data = request.json
    pn_query = data.get('product_number', '')
    if not pn_query:
        return jsonify({'status': 'error', 'message': '품번 없음.'}), 400
    
    try:
        settings_query = Setting.query.filter_by(brand_id=current_user.current_brand_id).all()
        brand_settings = {s.key: s.value for s in settings_query}

        search_term_cleaned = clean_string_upper(pn_query)
        search_like = f"%{search_term_cleaned}%"
        
        product = Product.query.options(
            selectinload(Product.variants)
        ).filter(
            Product.brand_id == current_user.current_brand_id,
            Product.product_number_cleaned.like(search_like)
        ).first()

        if product:
            variants = sorted(product.variants, key=lambda v: get_sort_key(v, brand_settings)) 
            colors = []
            if variants:
                colors_seen = set()
                for v in variants:
                    if v.color not in colors_seen:
                        colors.append(v.color)
                        colors_seen.add(v.color)
                        
            sizes = []
            if variants:
                sizes_seen = set()
                for v in variants:
                     if v.size not in sizes_seen:
                        sizes.append(v.size)
                        sizes_seen.add(v.size)
            
            return jsonify({
                'status': 'success',
                'product_name': product.product_name,
                'product_number': product.product_number,
                'colors': colors,
                'sizes': sizes
            })
        else:
            return jsonify({
                'status': 'error',
                'message': f"'{pn_query}'(으)로 시작하는 상품을 찾을 수 없습니다."
            }), 404
    except Exception as e:
        print(f"Find product details error: {e}")
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/order_product_search', methods=['POST'])
@login_required
def api_order_product_search():
    if not current_user.store_id:
        abort(403, description="이 API는 매장 계정만 사용할 수 있습니다.")
        
    data = request.json
    query = data.get('query', '')
    if not query:
        return jsonify({'status': 'error', 'message': '검색어 없음.'}), 400
    
    search_term_cleaned = clean_string_upper(query)
    search_like = f"%{search_term_cleaned}%"
    
    products = Product.query.filter(
        Product.brand_id == current_user.current_brand_id,
        or_(
            Product.product_number_cleaned.like(search_like),
            Product.product_name_cleaned.like(search_like),
            Product.product_name_choseong.like(search_like)
        )
    ).order_by(Product.product_name).limit(20).all()

    if products:
        results = [{
            'product_id': p.id,
            'product_number': p.product_number,
            'product_name': p.product_name
        } for p in products]
        return jsonify({'status': 'success', 'products': results})
    else:
        return jsonify({'status': 'error', 'message': f"'{query}'(으)로 검색된 상품이 없습니다."}), 404

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string