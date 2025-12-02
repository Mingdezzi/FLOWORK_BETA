import pandas as pd
import numpy as np
from openpyxl.utils import column_index_from_string
from flowork.utils import clean_string_upper, get_choseong, generate_barcode
import traceback
import json
import os
from flask import current_app
from flowork.models import db, Product, Variant, StoreStock, Setting, Brand

try:
    from flowork.services.transformer import transform_horizontal_to_vertical
except ImportError:
    transform_horizontal_to_vertical = None

def _get_column_indices_from_form(form, field_map, strict=True):
    column_map_indices = {}
    missing_fields = []
    
    for field_name, (form_key, is_required) in field_map.items():
        col_letter = form.get(form_key)
        
        if strict and is_required and not col_letter:
            missing_fields.append(field_name)
        
        if col_letter:
            try:
                column_map_indices[field_name] = column_index_from_string(col_letter) - 1
            except ValueError:
                column_map_indices[field_name] = None
        else:
            column_map_indices[field_name] = None

    if missing_fields:
        raise ValueError(f"다음 필수 항목의 열이 선택되지 않았습니다: {', '.join(missing_fields)}")
            
    return column_map_indices

def _read_excel_data_to_df(file_stream, column_map_indices):
    try:
        if hasattr(file_stream, 'seek'):
            file_stream.seek(0)
        df = pd.read_excel(file_stream, header=0)
    except Exception:
        if hasattr(file_stream, 'seek'):
            file_stream.seek(0)
        try:
            df = pd.read_csv(file_stream)
        except:
            return pd.DataFrame()

    if df.empty:
        return pd.DataFrame()

    selected_cols = {}
    total_cols = df.shape[1]
    for field_name, col_idx in column_map_indices.items():
        if col_idx is not None and 0 <= col_idx < total_cols:
            original_col_name = df.columns[col_idx]
            selected_cols[original_col_name] = field_name
    
    if not selected_cols:
        return pd.DataFrame()

    df_subset = df[list(selected_cols.keys())].rename(columns=selected_cols)
    
    for field in column_map_indices.keys():
        if field not in df_subset.columns:
            df_subset[field] = np.nan
            
    return df_subset

def _optimize_dataframe(df, brand_settings, upload_mode):
    if df.empty: return df

    required = ['product_number', 'color', 'size']
    df = df.dropna(subset=required)
    
    str_cols = ['product_number', 'product_name', 'color', 'size', 'item_category']
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().replace({'nan': None, 'None': None})

    num_cols = ['original_price', 'sale_price', 'release_year', 'hq_stock', 'store_stock']
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

    if 'original_price' in df.columns and 'sale_price' in df.columns:
        op = df['original_price']
        sp = df['sale_price']
        df['sale_price'] = np.where((op > 0) & (sp == 0), op, sp)
        df['original_price'] = np.where((sp > 0) & (op == 0), sp, op)

    df['product_number_cleaned'] = df['product_number'].apply(clean_string_upper)
    
    if 'color' in df.columns:
        df['color_cleaned'] = df['color'].apply(clean_string_upper)
    if 'size' in df.columns:
        df['size_cleaned'] = df['size'].apply(clean_string_upper)
    
    if 'product_name' in df.columns:
        df['product_name_cleaned'] = df['product_name'].apply(clean_string_upper)
        df['product_name_choseong'] = df['product_name'].apply(get_choseong)

    # [수정] 필수 데이터가 정제 후 빈 값이 된 경우 해당 행 삭제 (DB 에러 방지 핵심)
    # 예: color가 "-" 등이어서 clean_string_upper 후 ""가 된 경우 필터링
    for col in ['product_number_cleaned', 'color_cleaned', 'size_cleaned']:
        if col in df.columns:
            df = df[df[col] != '']

    if 'barcode' not in df.columns: df['barcode'] = None
    
    mask_no_barcode = df['barcode'].isna() | (df['barcode'] == '')
    if mask_no_barcode.any():
        df.loc[mask_no_barcode, 'barcode'] = df[mask_no_barcode].apply(
            lambda row: generate_barcode(row.to_dict(), brand_settings), axis=1
        )

    df = df.dropna(subset=['barcode'])
    df['barcode_cleaned'] = df['barcode'].apply(clean_string_upper)
    
    if 'is_favorite' not in df.columns:
        df['is_favorite'] = 0
    else:
        df['is_favorite'] = pd.to_numeric(df['is_favorite'], errors='coerce').fillna(0).astype(int)
    
    df = df.drop_duplicates(subset=['barcode_cleaned'], keep='last')

    return df

def verify_stock_excel(file_path, form, upload_mode):
    field_map = {'product_number': ('col_pn', True)}
    
    try:
        column_map_indices = _get_column_indices_from_form(form, field_map, strict=False)
        
        with open(file_path, 'rb') as f:
            df = _read_excel_data_to_df(f, column_map_indices)
        
        if df.empty:
            return {'status': 'success', 'suspicious_rows': []}

        df['_row_index'] = df.index + 2
        suspicious_rows = []
        
        for _, row in df.iterrows():
            pn = row.get('product_number')
            if pd.isna(pn) or str(pn).strip() == "":
                suspicious_rows.append({
                    'row_index': int(row['_row_index']), 
                    'preview': '(품번없음)', 
                    'reasons': '품번 누락'
                })
                
        return {'status': 'success', 'suspicious_rows': suspicious_rows[:100]}

    except Exception as e:
        return {'status': 'error', 'message': f"검증 중 오류: {e}"}

def parse_stock_excel(file_path, form, upload_mode, brand_id, excluded_row_indices=None):
    try:
        settings_query = Setting.query.filter_by(brand_id=brand_id).all()
        brand_settings = {s.key: s.value for s in settings_query}
        
        if 'SIZE_MAPPING' not in brand_settings or 'CATEGORY_MAPPING_RULE' not in brand_settings:
            try:
                brand = db.session.get(Brand, brand_id)
                if brand:
                    json_path = os.path.join(current_app.root_path, 'brands', f'{brand.brand_name}.json')
                    if os.path.exists(json_path):
                        with open(json_path, 'r', encoding='utf-8') as f:
                            file_config = json.load(f)
                            for k, v in file_config.items():
                                if k not in brand_settings:
                                    brand_settings[k] = json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else str(v)
            except Exception as e:
                print(f"Config file fallback failed: {e}")

        is_horizontal = form.get('is_horizontal') == 'on'

        field_map = {
            'product_number': ('col_pn', True),
            'color': ('col_color', True),
            'product_name': ('col_pname', False),
            'release_year': ('col_year', False),
            'item_category': ('col_category', False),
            'original_price': ('col_oprice', False),
            'sale_price': ('col_sprice', False),
            'is_favorite': ('col_favorite', False)
        }
        
        import_strategy = None

        if upload_mode == 'hq':
            if is_horizontal:
                import_strategy = 'horizontal_matrix'
            else:
                field_map['size'] = ('col_size', True)
                field_map['hq_stock'] = ('col_hq_stock', True)

        elif upload_mode == 'store':
            if is_horizontal:
                import_strategy = 'horizontal_matrix'
            else:
                field_map['size'] = ('col_size', True)
                field_map['store_stock'] = ('col_store_stock', True)
        
        elif upload_mode == 'db': 
             if is_horizontal:
                import_strategy = 'horizontal_matrix'
             else:
                field_map['size'] = ('col_size', True)
                field_map['hq_stock'] = ('col_hq_stock', False)

        column_map_indices = _get_column_indices_from_form(form, field_map, strict=False)

        with open(file_path, 'rb') as f:
            df = pd.DataFrame()
            if import_strategy == 'horizontal_matrix' and transform_horizontal_to_vertical:
                try:
                    size_conf = json.loads(brand_settings.get('SIZE_MAPPING', '{}'))
                    cat_conf = json.loads(brand_settings.get('CATEGORY_MAPPING_RULE', '{}'))
                    
                    df = transform_horizontal_to_vertical(f, size_conf, cat_conf, column_map_indices)
                    
                    if upload_mode == 'store' and 'hq_stock' in df.columns:
                        df.rename(columns={'hq_stock': 'store_stock'}, inplace=True)
                        
                except Exception as e:
                    return None, f"매트릭스 변환 오류: {e}"
            else:
                df = _read_excel_data_to_df(f, column_map_indices)
            
        if df.empty:
            return None, "처리할 데이터가 없습니다."
            
        if excluded_row_indices:
            if '_row_index' in df.columns:
                df = df[~df['_row_index'].isin(excluded_row_indices)]

        df = _optimize_dataframe(df, brand_settings, upload_mode)
        
        if df.empty: 
            return None, "유효한 데이터 없음 (필수 정보 누락 등)"

        return df.to_dict('records'), None

    except Exception as e:
        traceback.print_exc()
        return None, f"파싱 오류: {e}"

def export_db_to_excel(brand_id):
    import io
    import openpyxl
    from datetime import datetime
    try:
        query = db.session.query(
            Product.product_number, Product.product_name, Product.release_year, Product.item_category,
            Variant.barcode, Variant.color, Variant.size, Variant.original_price, Variant.sale_price, Variant.hq_quantity,
            Product.is_favorite
        ).join(Variant, Product.id == Variant.product_id).filter(Product.brand_id == brand_id)
        
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.append(["품번", "품명", "연도", "카테고리", "바코드", "컬러", "사이즈", "정상가", "판매가", "본사재고", "즐겨찾기"])
        
        for row in query.yield_per(1000):
            ws.append(list(row))
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, f"db_backup_{datetime.now().strftime('%Y%m%d')}.xlsx", None
    except Exception as e:
        print(f"Export Error: {e}")
        traceback.print_exc()
        return None, None, str(e)

def export_stock_check_excel(store_id, brand_id):
    import io
    import openpyxl
    from datetime import datetime
    try:
        variants = db.session.query(Variant).join(Product).filter(Product.brand_id == brand_id).all()
        stocks = db.session.query(StoreStock).filter_by(store_id=store_id).all()
        stock_map = {s.variant_id: s for s in stocks}
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["품번", "품명", "컬러", "사이즈", "바코드", "전산재고", "실사재고", "차이"])
        
        for v in variants:
            st = stock_map.get(v.id)
            qty = st.quantity if st else 0
            actual = st.actual_stock if st and st.actual_stock is not None else ''
            diff = (qty - actual) if isinstance(actual, int) else ''
            ws.append([v.product.product_number, v.product.product_name, v.color, v.size, v.barcode, qty, actual, diff])
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, f"stock_check_{datetime.now().strftime('%Y%m%d')}.xlsx", None
    except Exception as e:
        return None, None, str(e)