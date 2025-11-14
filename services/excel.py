import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from flask import flash
from flowork.models import db, Product, Variant, Store, StoreStock, Setting
from flowork.utils import clean_string_upper, get_choseong, generate_barcode
from sqlalchemy import exc, or_
from sqlalchemy.orm import selectinload, joinedload
import io
from datetime import datetime
import traceback
import re
import json

try:
    from flowork.services.transformer import transform_horizontal_to_vertical
except ImportError:
    transform_horizontal_to_vertical = None


def _get_column_indices_from_form(form, field_map):
    column_map_letters = {}
    missing_fields = []
    
    for field_name, (form_key, is_required) in field_map.items():
        col_letter = form.get(form_key)
        if is_required and not col_letter:
            missing_fields.append(field_name)
        column_map_letters[field_name] = col_letter

    if missing_fields:
        raise ValueError(f"필수 항목의 엑셀 열을 선택해야 합니다: {', '.join(missing_fields)}")

    column_map_indices = {}
    for field, letter in column_map_letters.items():
        if letter:
            column_map_indices[field] = column_index_from_string(letter) - 1
        else:
            column_map_indices[field] = None
            
    return column_map_indices


def _read_excel_data_by_indices(ws, column_map_indices):
    data = []
    max_col_idx = max(filter(None, column_map_indices.values())) + 1
    
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=max_col_idx)):
        item = {'_row_index': i + 2} 
        has_data = False
        
        for key, col_idx in column_map_indices.items():
            if col_idx is not None:
                cell_value = row[col_idx].value
                item[key] = cell_value
                if cell_value is not None and str(cell_value).strip() != "":
                    has_data = True
            else:
                item[key] = None
        
        if not has_data:
            continue
            
        data.append(item)
    return data


def verify_stock_excel(file_path, form, stock_type):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
    except Exception as e:
        return {'status': 'error', 'message': f'파일 읽기 오류: {e}'}

    field_map = {
        'product_number': ('col_pn', False),
        'product_name': ('col_pname', False),
        'original_price': ('col_oprice', False),
        'sale_price': ('col_sprice', False),
        'qty': ('col_store_stock' if stock_type == 'store' else 'col_hq_stock', False)
    }
    
    if 'barcode_col' in form:
        field_map = {
            'barcode': ('barcode_col', True),
            'qty': ('qty_col', True)
        }

    try:
        column_map_indices = _get_column_indices_from_form(form, field_map)
        raw_data = _read_excel_data_by_indices(ws, column_map_indices)
        
        suspicious_rows = []
        
        for item in raw_data:
            reasons = []
            row_idx = item['_row_index']
            
            pk = item.get('product_number') or item.get('barcode')
            if not pk or str(pk).strip() == "":
                reasons.append("식별값(품번/바코드) 누락")
            
            numeric_fields = ['original_price', 'sale_price', 'qty']
            for field in numeric_fields:
                val = item.get(field)
                if val is not None and str(val).strip() != "":
                    clean_val = str(val).replace(',', '').replace('.', '').strip()
                    if not clean_val.isdigit() and clean_val != '':
                        if not (clean_val.startswith('-') and clean_val[1:].isdigit()):
                             reasons.append(f"'{field}' 필드에 문자 포함 ('{val}')")

            if reasons:
                preview_str = f"{pk if pk else '(없음)'}"
                if item.get('product_name'):
                    preview_str += f" / {item['product_name']}"
                
                suspicious_rows.append({
                    'row_index': row_idx,
                    'preview': preview_str,
                    'reasons': ", ".join(reasons)
                })

        return {'status': 'success', 'suspicious_rows': suspicious_rows}

    except ValueError as ve:
        return {'status': 'error', 'message': str(ve)}
    except Exception as e:
        traceback.print_exc()
        return {'status': 'error', 'message': f"검증 중 오류: {e}"}


def import_excel_file(file, form, brand_id, progress_callback=None):
    if not file:
        return False, '파일이 없습니다.', 'error'

    BATCH_SIZE = 500
    
    try:
        settings_query = Setting.query.filter_by(brand_id=brand_id).all()
        brand_settings = {s.key: s.value for s in settings_query}
        
        import_strategy = brand_settings.get('IMPORT_STRATEGESY')
        
        field_map = {
            'product_number': ('col_pn', True),
            'product_name': ('col_pname', True),
            'release_year': ('col_year', True),
            'item_category': ('col_category', True),
            'color': ('col_color', True),
            'size': ('col_size', True),
            'original_price': ('col_oprice', False),
            'sale_price': ('col_sprice', False),
            'is_favorite': ('col_favorite', False),
            'hq_stock': ('col_hq_stock', False),
        }

        if import_strategy == 'horizontal_matrix':
            field_map['size'] = ('col_size', False)
            field_map['hq_stock'] = ('col_hq_stock', False)
        
        column_map_indices = _get_column_indices_from_form(form, field_map)

        data = []
        
        if import_strategy == 'horizontal_matrix':
            if transform_horizontal_to_vertical is None:
                return False, '서버에 pandas 라이브러리가 설치되지 않아 변환 기능을 사용할 수 없습니다.', 'error'
            
            size_mapping_json = brand_settings.get('SIZE_MAPPING', '{}')
            try:
                size_mapping_config = json.loads(size_mapping_json)
            except json.JSONDecodeError:
                return False, '브랜드 설정 오류: SIZE_MAPPING 형식이 올바르지 않습니다.', 'error'

            category_mapping_json = brand_settings.get('CATEGORY_MAPPING_RULE', '{}')
            try:
                category_mapping_config = json.loads(category_mapping_json)
            except json.JSONDecodeError:
                return False, '브랜드 설정 오류: CATEGORY_MAPPING_RULE 형식이 올바르지 않습니다.', 'error'

            try:
                data = transform_horizontal_to_vertical(
                    file, 
                    size_mapping_config, 
                    category_mapping_config,
                    column_map_indices 
                )
            except Exception as e:
                traceback.print_exc()
                return False, f'엑셀 변환 중 오류 발생: {e}', 'error'
                
        else:
            try:
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                data = _read_excel_data_by_indices(ws, column_map_indices)
            except Exception as e:
                 return False, f'엑셀 파일 읽기 오류: {e}', 'error'

        validated_data = []
        errors = []
        seen_barcodes = set()

        for i, item in enumerate(data):
            row_num = i + 2 
            
            try:
                item['barcode'] = generate_barcode(item, brand_settings)
                
                if not item.get('barcode'):
                    if import_strategy: continue 
                    errors.append(f"{row_num}행: 바코드 생성 실패")
                    continue
                
                item['barcode_cleaned'] = clean_string_upper(item['barcode'])
                
                item['product_number'] = str(item['product_number']).strip()
                item['product_name'] = str(item['product_name']).strip()
                item['color'] = str(item['color']).strip()
                item['size'] = str(item['size']).strip()
                item['original_price'] = int(item.get('original_price') or 0)
                item['sale_price'] = int(item.get('sale_price') or item['original_price'])
                item['release_year'] = int(item.get('release_year') or 0) if item.get('release_year') else None
                item['item_category'] = str(item['item_category']).strip() if item.get('item_category') else None
                item['is_favorite'] = 1 if item.get('is_favorite') in [True, 1, '1', 'Y', 'O'] else 0
                
                item['product_number_cleaned'] = clean_string_upper(item['product_number'])
                item['product_name_cleaned'] = clean_string_upper(item['product_name'])
                item['product_name_choseong'] = get_choseong(item['product_name'])
                item['color_cleaned'] = clean_string_upper(item['color'])
                item['size_cleaned'] = clean_string_upper(item['size'])
                
                item['hq_stock'] = int(item.get('hq_stock') or 0) 

            except (ValueError, TypeError) as e:
                errors.append(f"{row_num}행 데이터 오류: {e}")
                continue

            if item['barcode_cleaned'] in seen_barcodes:
                if not import_strategy:
                    errors.append(f"{row_num}행: 바코드 중복 ({item['barcode']})")
                continue
            seen_barcodes.add(item['barcode_cleaned'])
            
            validated_data.append(item)
            
        if errors and not import_strategy:
             return False, f"검증 오류 (최대 5개): {', '.join(errors[:5])}", 'error'

        store_ids_to_delete = db.session.query(Store.id).filter_by(brand_id=brand_id)
        db.session.query(StoreStock).filter(StoreStock.store_id.in_(store_ids_to_delete)).delete(synchronize_session=False)
        product_ids_to_delete = db.session.query(Product.id).filter_by(brand_id=brand_id)
        db.session.query(Variant).filter(Variant.product_id.in_(product_ids_to_delete)).delete(synchronize_session=False)
        db.session.query(Product).filter_by(brand_id=brand_id).delete(synchronize_session=False)
        
        db.session.commit()

        products_map = {}
        total_products_created = 0
        total_variants_created = 0
        
        total_items = len(validated_data)

        for i in range(0, len(validated_data), BATCH_SIZE):
            if progress_callback:
                progress_callback(i, total_items)

            batch_data = validated_data[i:i+BATCH_SIZE]
            products_to_add_batch = []
            variants_to_add_batch = []
            
            for item in batch_data:
                pn_key = item['product_number_cleaned']
                if pn_key not in products_map:
                    product = Product(
                        brand_id=brand_id,
                        product_number=item['product_number'],
                        product_name=item['product_name'],
                        release_year=item['release_year'],
                        item_category=item['item_category'],
                        is_favorite=item['is_favorite'],
                        product_number_cleaned=item['product_number_cleaned'],
                        product_name_cleaned=item['product_name_cleaned'],
                        product_name_choseong=item['product_name_choseong']
                    )
                    products_map[pn_key] = product
                    products_to_add_batch.append(product)

            if products_to_add_batch:
                db.session.add_all(products_to_add_batch)
                total_products_created += len(products_to_add_batch)

            try:
                db.session.flush()
            except Exception as e:
                db.session.rollback()
                print(f"Error flushing products batch: {e}")
                return False, f"DB 저장 실패 (Product Flush): {e}", 'error'
            
            for item in batch_data:
                pn_key = item['product_number_cleaned']
                product = products_map.get(pn_key)
                
                if not product or not product.id:
                    continue
                    
                variant = Variant(
                    product_id=product.id, 
                    barcode=item['barcode'],
                    color=item['color'],
                    size=item['size'],
                    original_price=item['original_price'],
                    sale_price=item['sale_price'],
                    hq_quantity=item['hq_stock'],
                    barcode_cleaned=item['barcode_cleaned'],
                    color_cleaned=clean_string_upper(item['color']),
                    size_cleaned=clean_string_upper(item['size'])
                )
                variants_to_add_batch.append(variant)

            if variants_to_add_batch:
                db.session.bulk_save_objects(variants_to_add_batch)
                total_variants_created += len(variants_to_add_batch)
            
            db.session.commit()
        
        if progress_callback:
            progress_callback(total_items, total_items)
        
        return True, f'업로드 완료. (상품 {total_products_created}개, 옵션 {total_variants_created}개)', 'success'

    except ValueError as ve:
        return False, str(ve), 'error'
    except exc.IntegrityError as e:
        db.session.rollback()
        print(f"DB Import IntegrityError: {e}")
        traceback.print_exc()
        return False, f"DB 저장 실패 (데이터 중복 오류): {e.orig}", 'error'
    except Exception as e:
        db.session.rollback()
        print(f"DB Import Error: {e}")
        traceback.print_exc()
        return False, f"엑셀 처리 중 알 수 없는 오류 발생: {e}", 'error'


def process_stock_upsert_excel(file_path, form, stock_type, brand_id, target_store_id=None, progress_callback=None, excluded_row_indices=None, allow_create=True):
    try:
        settings_query = Setting.query.filter_by(brand_id=brand_id).all()
        brand_settings = {s.key: s.value for s in settings_query}
        import_strategy = brand_settings.get('IMPORT_STRATEGY')

        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            return 0, 0, f'엑셀 파일을 여는 중 오류 발생: {e}', 'error'

    except Exception as e:
        return 0, 0, f'설정 로드 중 오류 발생: {e}', 'error'

    if stock_type == 'store' and not target_store_id:
        return 0, 0, '매장 재고를 업데이트하려면 대상 매장 ID가 필요합니다.', 'error'

    try:
        field_map = {
            'product_number': ('col_pn', True),
            'product_name': ('col_pname', True),
            'color': ('col_color', True),
            'size': ('col_size', True),
            'original_price': ('col_oprice', True),
            'sale_price': ('col_sprice', True),
            'hq_stock': ('col_hq_stock', stock_type == 'hq'),
            'store_stock': ('col_store_stock', stock_type == 'store'),
            'release_year': ('col_year', False),
            'item_category': ('col_category', False),
        }
        
        if import_strategy == 'horizontal_matrix':
            field_map['size'] = ('col_size', False)
            field_map['hq_stock'] = ('col_hq_stock', False)
            field_map['store_stock'] = ('col_store_stock', False)

        active_field_map = {}
        if stock_type == 'hq':
            for k, v in field_map.items():
                if k != 'store_stock': active_field_map[k] = v
        else:
            for k, v in field_map.items():
                if k not in ['hq_stock', 'release_year', 'item_category']: active_field_map[k] = v

        column_map_indices = _get_column_indices_from_form(form, active_field_map)
        items_to_process = []

        if import_strategy == 'horizontal_matrix':
            if transform_horizontal_to_vertical is None:
                return 0, 0, '서버에 pandas 라이브러리가 없어 변환 기능을 사용할 수 없습니다.', 'error'
            
            try:
                size_mapping_config = json.loads(brand_settings.get('SIZE_MAPPING', '{}'))
                category_mapping_config = json.loads(brand_settings.get('CATEGORY_MAPPING_RULE', '{}'))
            except json.JSONDecodeError:
                return 0, 0, '브랜드 설정(SIZE_MAPPING 등) 형식이 올바르지 않습니다.', 'error'

            try:
                with open(file_path, 'rb') as f:
                    items_to_process = transform_horizontal_to_vertical(
                        f, 
                        size_mapping_config, 
                        category_mapping_config,
                        column_map_indices 
                    )
                
                for item in items_to_process:
                    qty = item.get('hq_stock') or item.get('quantity') or 0
                    if stock_type == 'store':
                        item['store_stock'] = qty
                    else:
                        item['hq_stock'] = qty

            except Exception as e:
                traceback.print_exc()
                return 0, 0, f'엑셀 변환 중 오류 발생: {e}', 'error'
        
        else:
            ws = wb.active
            items_to_process = _read_excel_data_by_indices(ws, column_map_indices)

        excluded_set = set(excluded_row_indices) if excluded_row_indices else set()
        if excluded_set:
            items_to_process = [it for it in items_to_process if it.get('_row_index') not in excluded_set]

        total_items = len(items_to_process)
        if total_items == 0:
            return 0, 0, '엑셀에서 유효한 데이터를 찾을 수 없습니다.', 'warning'

        if progress_callback:
            progress_callback(0, total_items)

        pn_cleaned_list = list(set(clean_string_upper(item['product_number']) for item in items_to_process if item.get('product_number')))
        
        products_in_db = Product.query.filter(
            Product.brand_id == brand_id,
            Product.product_number_cleaned.in_(pn_cleaned_list)
        ).options(selectinload(Product.variants)).all()
        
        product_map = {p.product_number_cleaned: p for p in products_in_db}
        variant_map = {}
        for p in products_in_db:
            for v in p.variants:
                variant_map[v.barcode_cleaned] = v

        store_stock_map = {}
        if stock_type == 'store':
            variant_ids_in_db = [v.id for v in variant_map.values()]
            if variant_ids_in_db:
                existing_stock = db.session.query(StoreStock).filter(
                    StoreStock.store_id == target_store_id,
                    StoreStock.variant_id.in_(variant_ids_in_db)
                ).all()
                store_stock_map = {s.variant_id: s for s in existing_stock}

        created_product_count = 0
        created_variant_count = 0
        updated_variant_price_count = 0
        updated_hq_stock_count = 0
        created_store_stock_count = 0
        updated_store_stock_count = 0
        
        new_products_to_add = []
        new_variants_to_add = []
        variants_to_update_stock = []
        items_for_store_stock = []

        for idx, item in enumerate(items_to_process):
            if progress_callback and idx % 50 == 0:
                progress_callback(idx, total_items)

            try:
                pn = str(item.get('product_number', '')).strip()
                pname = str(item.get('product_name', '')).strip()
                color = str(item.get('color', '')).strip()
                size = str(item.get('size', '')).strip()
                
                if not pn or not color or not size:
                    continue

                release_year = None
                if item.get('release_year'):
                    raw_year = str(item.get('release_year', '')).replace(' ', '').replace('년', '').strip()
                    if raw_year:
                        try:
                            release_year = int(float(raw_year))
                        except ValueError:
                            pass

                barcode_item = {'product_number': pn, 'color': color, 'size': size}
                barcode = generate_barcode(barcode_item, brand_settings)
                if not barcode:
                    print(f"Skipping row (barcode gen failed): {item}")
                    continue
                
                barcode_cleaned = clean_string_upper(barcode)
                pn_cleaned = clean_string_upper(pn)
                
                product = product_map.get(pn_cleaned)
                if not product:
                    if not allow_create:
                        continue

                    product = Product(
                        brand_id=brand_id,
                        product_number=pn,
                        product_name=pname,
                        product_number_cleaned=pn_cleaned,
                        product_name_cleaned=clean_string_upper(pname),
                        product_name_choseong=get_choseong(pname),
                        release_year=release_year, 
                        item_category=str(item.get('item_category', '')).strip() if item.get('item_category') else None
                    )
                    product_map[pn_cleaned] = product
                    new_products_to_add.append(product)
                    created_product_count += 1
                
                variant = variant_map.get(barcode_cleaned)
                original_price = int(item.get('original_price') or 0)
                sale_price = int(item.get('sale_price') or original_price)

                if not variant:
                    if not allow_create:
                        continue

                    variant = Variant(
                        product=product,
                        barcode=barcode,
                        color=color,
                        size=size,
                        original_price=original_price,
                        sale_price=sale_price,
                        hq_quantity=0,
                        barcode_cleaned=barcode_cleaned,
                        color_cleaned=clean_string_upper(color),
                        size_cleaned=clean_string_upper(size)
                    )
                    variant_map[barcode_cleaned] = variant
                    new_variants_to_add.append(variant)
                    created_variant_count += 1
                else:
                    if original_price > 0: variant.original_price = original_price
                    if sale_price > 0: variant.sale_price = sale_price
                    updated_variant_price_count += 1
                
                if stock_type == 'hq':
                    qty = int(item.get('hq_stock') or 0)
                    variants_to_update_stock.append((variant, qty))
                elif stock_type == 'store':
                    qty = int(item.get('store_stock') or 0)
                    items_for_store_stock.append((variant, qty))
            
            except Exception as e:
                print(f"Skipping row (data error: {e}): {item}")
                continue
        
        if progress_callback:
            progress_callback(int(total_items * 0.9), total_items)

        if new_products_to_add:
            db.session.add_all(new_products_to_add)
        if new_variants_to_add:
            db.session.add_all(new_variants_to_add)
            
        if new_products_to_add or new_variants_to_add:
            try:
                db.session.flush()
            except exc.IntegrityError as e:
                db.session.rollback()
                return 0, 0, f"DB 저장 실패 (중복 등): {e.orig}", 'error'

        if stock_type == 'hq':
            for variant, hq_qty in variants_to_update_stock:
                variant.hq_quantity = hq_qty
                updated_hq_stock_count += 1
        
        elif stock_type == 'store':
            new_store_stock_entries = []

            variant_qty_map = {}
            for variant, store_qty in items_for_store_stock:
                if variant.id:
                    variant_qty_map[variant.id] = store_qty 
            
            updated_store_stock_count_in_batch = 0

            for variant_id, store_qty in variant_qty_map.items():
                stock_entry = store_stock_map.get(variant_id)
                
                if stock_entry:
                    stock_entry.quantity = store_qty
                    updated_store_stock_count_in_batch += 1
                else:
                    new_stock = StoreStock(
                        store_id=target_store_id,
                        variant_id=variant_id,
                        quantity=store_qty,
                        actual_stock=None
                    )
                    new_store_stock_entries.append(new_stock)
                    created_store_stock_count += 1
            
            updated_store_stock_count += updated_store_stock_count_in_batch

            if new_store_stock_entries:
                db.session.add_all(new_store_stock_entries)

        db.session.commit()
        
        if progress_callback:
            progress_callback(total_items, total_items)
        
        if stock_type == 'hq':
             msg = f"본사재고 UPSERT 완료. (신규 상품: {created_product_count} / 신규 옵션: {created_variant_count}) (옵션 가격 수정: {updated_variant_price_count} / 본사 재고 수정: {updated_hq_stock_count})"
             total_processed = updated_hq_stock_count
        else:
             msg = f"매장재고 UPSERT 완료. (신규 상품: {created_product_count} / 신규 옵션: {created_variant_count}) (옵션 가격 수정: {updated_variant_price_count}) (매장 재고 생성: {created_store_stock_count} / 매장 재고 수정: {updated_store_stock_count})"
             total_processed = created_store_stock_count + updated_store_stock_count
        
        total_created = created_product_count + created_variant_count
        
        return total_processed, total_created, msg, 'success'

    except ValueError as ve:
        return 0, 0, str(ve), 'error'
    except Exception as e:
        db.session.rollback()
        print(f"Stock UPSERT Error: {e}")
        traceback.print_exc()
        return 0, 0, f'엑셀 처리 중 오류 발생: {e}', 'error'


def _process_stock_update_excel(file, form, stock_type, brand_id, target_store_id):
    try:
        if isinstance(file, str):
            wb = openpyxl.load_workbook(file, data_only=True)
        else:
            if not file or not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
                return 0, 0, '엑셀 파일(.xlsx, .xls)을 업로드하세요.', 'error'
            wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        return 0, 0, f'엑셀 파일 열기 오류: {e}', 'error'

    if stock_type == 'store' and not target_store_id:
        return 0, 0, '매장 재고를 업데이트하려면 대상 매장 ID(target_store_id)가 필요합니다.', 'error'

    try:
        field_map = {
            'barcode': ('barcode_col', True),
            'qty': ('qty_col', True),
        }
        column_map_indices = _get_column_indices_from_form(form, field_map)

        ws = wb.active
        items = _read_excel_data_by_indices(ws, column_map_indices)
        
        barcode_map = {}
        for item in items:
            try:
                barcode = str(item['barcode']).strip() if item['barcode'] else None
                qty = int(item['qty']) if item['qty'] is not None else 0
                if barcode:
                    barcode_map[clean_string_upper(barcode)] = qty
            except (ValueError, TypeError):
                continue 

        if not barcode_map:
            return 0, 0, '엑셀에서 유효한 재고 데이터를 찾을 수 없습니다.', 'warning'

        variants = db.session.query(Variant).join(Product).filter(
            Product.brand_id == brand_id,
            Variant.barcode_cleaned.in_(barcode_map.keys())
        ).all() 
        
        if not variants:
            return 0, 0, 'DB에 일치하는 상품(바코드)이 없습니다. (신규 상품은 생성되지 않습니다)', 'error'

        updated_count = 0
        added_count = 0
        
        if stock_type == 'store':
            variant_id_map = {v.id for v in variants}
            
            existing_stock_query = db.session.query(StoreStock).filter(
                StoreStock.store_id == target_store_id,
                StoreStock.variant_id.in_(variant_id_map)
            ).all()
            
            stock_map = {s.variant_id: s for s in existing_stock_query}
            variant_barcode_to_id_map = {v.barcode_cleaned: v.id for v in variants}
            
            new_stock_entries = []

            for barcode_cleaned, variant_id in variant_barcode_to_id_map.items():
                new_qty = barcode_map.get(barcode_cleaned, 0)
                
                if variant_id in stock_map:
                    stock_entry = stock_map[variant_id]
                    stock_entry.quantity = new_qty
                    updated_count += 1
                else:
                    new_stock = StoreStock(
                        store_id=target_store_id,
                        variant_id=variant_id,
                        quantity=new_qty,
                        actual_stock=None
                    )
                    new_stock_entries.append(new_stock)
                    added_count += 1
            
            if new_stock_entries:
                db.session.add_all(new_stock_entries)
            
            stock_name = '매장 재고'
            
        elif stock_type == 'hq':
            for variant in variants:
                new_qty = barcode_map.get(variant.barcode_cleaned, 0)
                variant.hq_quantity = new_qty
                updated_count += 1
            
            stock_name = '본사 재고'
        
        else:
            return 0, 0, f"알 수 없는 stock_type: {stock_type}", 'error'

        db.session.commit()
        
        message = f"엑셀 {stock_name} 업데이트 완료 (바코드 2필드 기준). (신규 {added_count}건, 업데이트 {updated_count}건)"
        return updated_count, added_count, message, 'success'

    except ValueError as ve:
        return 0, 0, str(ve), 'error'
    except Exception as e:
        db.session.rollback()
        print(f"Stock Excel Update (Barcode) Error: {e}")
        traceback.print_exc()
        return 0, 0, f'엑셀 처리 중 오류 발생: {e}', 'error'


def export_db_to_excel(brand_id):
    try:
        products_variants_query = db.session.query(
            Product.product_number,
            Product.product_name,
            Product.release_year,
            Product.item_category,
            Product.is_favorite,
            Variant.barcode,
            Variant.color,
            Variant.size,
            Variant.original_price,
            Variant.sale_price,
            Variant.hq_quantity,
        ).join(Variant, Product.id == Variant.product_id).filter(
            Product.brand_id == brand_id
        ).order_by(Product.product_number, Variant.id).execution_options(yield_per=100)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products_Variants_Backup"

        headers = [
            "품번", "품명", "연도", "카테고리", 
            "바코드", "컬러", "사이즈", "정상가", "판매가",
            "본사재고", "즐겨찾기"
        ]
        ws.append(headers)
        
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        is_empty = True
        for row in products_variants_query:
            if is_empty:
                 is_empty = False
                 
            product_number, product_name, release_year, item_category, is_favorite, barcode, color, size, original_price, sale_price, hq_quantity = row
            
            data_row = [
                product_number,
                product_name,
                release_year,
                item_category,
                barcode,
                color,
                size,
                original_price,
                sale_price,
                hq_quantity,
                is_favorite
            ]
            ws.append(data_row)
        
        if is_empty:
             return None, None, "백업할 상품 데이터가 없습니다."

        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        today_str = datetime.now().strftime('%Y%m%d')
        download_name = f'flowork_db_backup_{today_str}.xlsx'
        
        return output, download_name, None

    except Exception as e:
        db.session.rollback()
        print(f"DB Export Error: {e}")
        traceback.print_exc()
        return None, None, f"엑셀 백업 중 오류 발생: {e}"


def export_stock_check_excel(current_store_id, current_brand_id):
    try:
        all_variants_in_brand = db.session.query(Variant).join(Product).filter(
            Product.brand_id == current_brand_id
        ).options(
            joinedload(Variant.product)
        ).order_by(
            Product.product_number, Variant.color, Variant.size
        ).all()

        if not all_variants_in_brand:
            return None, None, "엑셀로 출력할 상품 데이터가 없습니다."
            
        all_variant_ids = [v.id for v in all_variants_in_brand]

        current_store_stock_query = db.session.query(StoreStock).filter(
            StoreStock.store_id == current_store_id,
            StoreStock.variant_id.in_(all_variant_ids)
        ).all()
        my_stock_map = {s.variant_id: s for s in current_store_stock_query}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "재고 실사 시트"

        headers = [
            "품번", "품명", "컬러", "사이즈", "바코드", 
            "현재고", "본사재고", "실사재고", "재고차이"
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        
        for i, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            if i <= 5: ws.column_dimensions[get_column_letter(i)].width = 18
            else: ws.column_dimensions[get_column_letter(i)].width = 12

        for row_num, variant in enumerate(all_variants_in_brand, 2):
            product = variant.product
            
            my_stock = my_stock_map.get(variant.id)
            
            my_qty = my_stock.quantity if my_stock else 0
            hq_qty = variant.hq_quantity
            actual_qty = my_stock.actual_stock if (my_stock and my_stock.actual_stock is not None) else ''
            
            diff = ''
            if isinstance(actual_qty, int):
                diff = my_qty - actual_qty

            ws.append([
                product.product_number,
                product.product_name,
                variant.color,
                variant.size,
                f" {variant.barcode}",
                my_qty,
                hq_qty,
                actual_qty,
                diff
            ])
            
            for col_idx in range(6, 10):
                ws.cell(row=row_num, column=col_idx).alignment = center_align


        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        today_str = datetime.now().strftime('%Y%m%d')
        download_name = f'flowork_stock_check_{today_str}.xlsx'
        
        return output, download_name, None

    except Exception as e:
        print(f"Stock Check Export Error: {e}")
        traceback.print_exc()
        return None, None, f"재고 실사 엑셀 생성 중 오류 발생: {e}"